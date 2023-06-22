import os

def create_column_if_not_exists(ws, statuscolumnheader):
    for col in ws.iter_cols():
        if statuscolumnheader.lower() == col[0].value.lower():
            return ws, col[0].column-1
    ws.insert_cols(ws.max_column+1)
    statuscell = ws.cell(row=1, column=ws.max_column+1)
    statuscell.value = statuscolumnheader
    return ws, statuscell.column-1

def create_columns_if_not_exists(ws, columns):
    """
    Creates the columns if they do not exist.

    parameters:
        `ws` (object): Worksheet object.
        `columns` (list): List of column headers.

    Returns:
        ws (Worksheet): Worksheet object.
        colss (dict): Dictionary of column names and their respective column numbers.
    """
    colss = {}
    for column in columns:
        for col in ws.iter_cols():
            if column.lower() == col[0].value.lower():
                colss[column] = col[0].column-1
                break
        else:
            ws.insert_cols(ws.max_column+1)
            statuscell = ws.cell(row=1, column=ws.max_column+1)
            statuscell.value = column
            colss[column] = statuscell.column-1
    return ws, colss

def update_excel_value(filename, criteriacolumnheader, criteriacellvalue, statuscolumnheader, statusvalue, remarkscolumnheader, remarksvalue):
    """
    Updates the excel file with the given values.

    Returns:
        True if the value is updated.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    # ws, statuscolumn = create_column_if_not_exists(ws, statuscolumnheader)
    ws, statuscolumn = create_columns_if_not_exists(ws, [statuscolumnheader, remarkscolumnheader])
    for col in ws.iter_cols():
        if criteriacolumnheader == col[0].value:
            for row in ws.iter_rows():
                if criteriacellvalue == str(row[col[0].column-1].value):
                    # row[statuscolumn].value = statusvalue
                    row[statuscolumn[statuscolumnheader]].value = statusvalue
                    row[statuscolumn[remarkscolumnheader]].value = remarksvalue
                    wb.save(filename)
                    return True


def kill_task_process(processname: str):
    """
    Kills the `processname` process if it is running.

    Parameters:
        `processname` (string): Name of the process.

    Returns:
        None

    Examples:
        | `Kill Task Process` | chrome |
        | `Kill Task Process` | excel  |
    """
    import os
    # os.system("taskkill /im " + processname + ".exe /f /t")
    if processname.lower() in os.popen("tasklist").read().lower():
        os.system("taskkill /im " + processname + ".exe /f")





def get_init_details():

    relpath = os.getcwd()

    # print(os.getcwd())

    #df_from_excel = pd.read_excel(relpath+"\input\Automation Test Case.xlsx", sheet_name='TestData').to_dict(orient='records')

    return(relpath)