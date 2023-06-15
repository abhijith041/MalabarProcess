import pandas as pd
from sqlalchemy import create_engine
import pymysql 
import pyodbc
from openpyxl import load_workbook

import configVariables



configVariables.dbConnectionString

#fetching legal entity, cost center name... from excel
def read_horizontaldata(inputpath):

    """
    This function is used to read the constant data of each input sheet that is the data that een in the top of all input sheets
    get all data and return that for further process

    """

    df = pd.read_excel(inputpath, header=None, skiprows=2)

    # Transpose the dataframe
    df = df.T

    # Set the first column as the header
    df.columns = df.iloc[0]

    # Remove the first row (header row)
    df = df[1:]

    # Reset the index
    df.reset_index(drop=True, inplace=True)

    # Accessing values by column name
    legal_entity_code = df['Legal Entity Code'][1]
    cost_center_code = df['Cost Center Code'][1]
    emp_code = df['Emp Code'][1]
    interior_code = df['Interior Code'][1]
    work_code = df['Work Code'][1]

       # Accessing values by column name
    legal_entity_code = df['Legal Entity Code'][1] if not pd.isna(df['Legal Entity Code'][1]) else 'Not found'
    cost_center_code = df['Cost Center Code'][1] if not pd.isna(df['Cost Center Code'][1]) else 'Not found'
    emp_code = df['Emp Code'][1] if not pd.isna(df['Emp Code'][1]) else 'Not found'
    interior_code = df['Interior Code'][1] if not pd.isna(df['Interior Code'][1]) else 'Not found'
    work_code = df['Work Code'][1] if not pd.isna(df['Work Code'][1]) else 'Not found'


    return  legal_entity_code, cost_center_code, emp_code, interior_code, work_code

 


# def create_columns(filename):
#         try:
#             wb = load_workbook(filename)
#             sheet = wb.active
#             max_column_index = sheet.max_column
#             num_cols = 1
#             index_position = max_column_index + 1
#             column_heading = 'duplication'
#             column_exists = any(column_heading == sheet.cell(row=8, column=j).value for j in range(1, max_column_index + 1))  
#             if not column_exists:
#                 sheet.cell(row=8, column=index_position, value=column_heading)
#             wb.save(filename)
#             print("Column creation successful.")
#         except Exception as e:
#             print("An error occurred:", str(e))
def create_columns(filename):      

    wb = load_workbook(filename)
    sheet = wb.active
    max_column_index = sheet.max_column
    num_cols = 3
    index_position = max_column_index + 1
    column_headings = ['duplication', 'status', 'comments']  

    for i in range(num_cols):
        column_heading = column_headings[i]
        column_exists = any(column_heading == sheet.cell(row=8, column=j).value for j in range(1, max_column_index + 1))       

        if not column_exists:
            cell = sheet.cell(row=8, column=index_position + i, value=column_heading)
    #sheet.insert_cols(idx=index_position, amount=num_cols)
    wb.save(filename)



#create_columns('C:/Users/Q0037/Documents/Malabar/VoucherVerificationProcess/InputFolder/Nehamathew@gmail.com/input.xlsx')

def duplicate_check_database(voucher_legal_PK, engine_str):

    try:
        engine = create_engine(configVariables.dbConnectionString)
        query = f"SELECT voucher_legalentity FROM project_petty_table WHERE voucher_legalentity = '{voucher_legal_PK}'"
        df = pd.read_sql(query, con=engine)
        
        if len(df) > 0:
            duplication = True
            # status = df.iloc[0]['status']
            # status = status.lower()
            
            # if status == 'valid':
            #     #duplicated and valid status data
            #     duplication = "Duplicated"
            # else:
            #     #duplicated but invalid status data
            #     duplication = "InvalidDuplicated"
            return  duplication
        else:
            duplication = False
        
            return duplication
    except Exception as e:      
      print(str(e))
      duplication = False
      return    duplication
 
 

def update_duplication_toExcel(filename, is_duplicated, voucher_no):
    try:

        driver = "Microsoft Excel Driver (*.xls, *.xlsx, *.xlsm, *.xlsb)"
        conn_str = f"Driver={{{driver}}};DBQ={filename};readOnly=0;"

        # Connect to Excel file using OLE DB
        cnxn = pyodbc.connect(conn_str, autocommit=True)

        # Open a cursor to execute SQL commands
        cursor = cnxn.cursor()

        # Update a cell in the worksheet
        worksheet_name = "Sheet1"       

        # Define the SQL statement to update the cell
        #update_sql = f"UPDATE [{worksheet_name}$A8:J] SET duplication = ? WHERE voucher_number = ?"
        update_sql = f"UPDATE [{worksheet_name}$A8:Z] SET duplication = ? WHERE voucher_number = ?"

        # Execute the SQL statement to update the cell
        cursor.execute(update_sql, [is_duplicated, voucher_no])

        # Commit the changes
        cnxn.commit()

        cursor.close()
        cnxn.close()
        print("Update successful.")
    except Exception as e:
        print("An error occurred:", str(e))




def duplication_checking_process(filepath,LE):
    """
    this is the main function used to check for duplication and inside from this function another functions were called those are written just above 
    from here.
    LE = legal entity name (eg: MGPP)
    """
    try:
    
        df = pd.read_excel(filepath, skiprows=7)
        df = df[df['credit'].isnull()]
        

        for index, row in df.iterrows():
            voucher_no = row['voucher_number']      
            vr_le = LE + '_' + str(voucher_no)
            print(vr_le)
            #Here checking the duplication occured in database calling another function
            # duplication =  duplicate_check_database(vr_le,'mysql+mysqlconnector://root:root@127.0.0.1:3306/malabarprocessdb')
            duplication =  duplicate_check_database(vr_le,configVariables.dbConnectionString)
            if duplication:
                status = 'Dupliacate data'
                #updating the duplcation status to excel 
                update_duplication_toExcel(filepath,status,voucher_no)
            else:
                status = 'Not duplicate'
                #updating the duplcation status to excel 
                update_duplication_toExcel(filepath,status,voucher_no)
    except Exception as e:
         print(str(e))
         


    #return  status
#duplication_process('C:/Users/Q0037/Documents/Malabar/VoucherVerificationProcess/InputFolder/Nehamathew@gmail.com/input.xlsx', 'MGPP')


def upload_input_values_DB(legal_entity_code,cost_center_code, emp_code, interior_code, work_code, inputpath,engine_str,mail_id):
    """
    Uploading the input sheet values to the database
    Also, update the status column value to New during the uploading time and this new status is need to fetch unprocessed data from DB
    Legal entity, cost center code, interior code, work code,emp code will be adding to each row items
    there may chance to empty voucher columns occur for credit row such case bot remove that column 
    if any of the voucher number is empty then we get that message through 'voucherNo_empty' variable if it's invalid that means there are rows with empty voucher number
    
    """
    try:
        excel_data = pd.read_excel(inputpath, skiprows=7)
        excel_data['voucher_legalentity'] = legal_entity_code +'_'+ excel_data['voucher_number'].astype(str)

       # Remove rows where credit column is not NaN
        excel_data = excel_data[excel_data['credit'].isnull()]

          # Filter rows where the value in the "duplicate" column is "No"
        excel_data = excel_data[excel_data['duplication'] == 'Not duplicate']

        current_row_count   =   len(excel_data)

        # checking current dataframe's length ,if length is 0 then goes to else case.
        if current_row_count != 0:

            noDataInDF =False
            print("there is data after removing duplicated items from dataframe.")

            # Check for NaN values in voucher number column
            if excel_data['voucher_number'].isnull().any():

                excel_data = excel_data.dropna(subset=['voucher_number'])
                # Assign "invalid" to voucher number column for NaN values
                voucherNo_empty = 'invalid'
            else:
                voucherNo_empty = None
            
            # Add "columns" column with "constant" values
            excel_data['status'] = 'New'
            excel_data['cost_center_code'] = cost_center_code
            excel_data['emp_code'] = emp_code
            excel_data['interior_code'] = interior_code
            excel_data['work_code'] = work_code
            excel_data['mail_id'] = mail_id
    
            #delete the duplication column from inputsheet
            excel_data.drop('duplication', inplace=True, axis=1)
            
            #engine = create_engine('mysql+pymysql://root:Password.123@127.0.0.1:3306/malabarprocessdb')
            engine = create_engine(configVariables.dbConnectionString)

            # Convert DataFrame to a list of dictionaries
            data = excel_data.to_dict(orient='records')
            
            # Format the date column
            excel_data['date'] = pd.to_datetime(excel_data['date'], format='%d/%m/%Y', errors='coerce')

            # Insert the data using pandas to_sql
            excel_data.to_sql("project_petty_table", con=engine, if_exists='append', index=False)
            
            # return noDataInDf =  False, dbupload = True, and voucherNo_empty 
            return noDataInDF, True, voucherNo_empty  # Data upload success


        else:
            noDataInDF =True
            print("no data found after removing duplicate items")
            voucherNo_empty = None
            
            # return noDataInDf =  True, dbupload = False, and voucherNo_empty
            return noDataInDF, False, voucherNo_empty

        # # Check for NaN values in voucher number column
        # if excel_data['voucher_number'].isnull().any():

        #     excel_data = excel_data.dropna(subset=['voucher_number'])
        #     # Assign "invalid" to voucher number column for NaN values
        #     voucherNo_empty = 'invalid'
        # else:
        #     voucherNo_empty = None
        
        # # Add "columns" column with "constant" values
        # excel_data['status'] = 'New'
        # excel_data['cost_center_code'] = cost_center_code
        # excel_data['emp_code'] = emp_code
        # excel_data['interior_code'] = interior_code
        # excel_data['work_code'] = work_code
        # excel_data['mail_id'] = mail_id
 
        # #delete the duplication column from inputsheet
        # excel_data.drop('duplication', inplace=True, axis=1)
        
        # #engine = create_engine('mysql+pymysql://root:Password.123@127.0.0.1:3306/malabarprocessdb')
        # engine = create_engine(configVariables.dbConnectionString)

        # # Convert DataFrame to a list of dictionaries
        # data = excel_data.to_dict(orient='records')
        
        # # Format the date column
        # excel_data['date'] = pd.to_datetime(excel_data['date'], format='%d/%m/%Y', errors='coerce')

        # # Insert the data using pandas to_sql
        # excel_data.to_sql("project_petty_table", con=engine, if_exists='append', index=False)
        
        # return True, voucherNo_empty  # Data upload success
        
    except pd.io.sql.DatabaseError as e:
        if 'Duplicate entry' in str(e):
            print('Error: Duplicate entry found. Please check for duplicate data.')
        elif 'for key' in str(e):
            print('Error: Primary key constraint violation. Please check the primary key column.')
        else:
            print(f'Error: {e}')

        # returning values  noDataInDf =  False, dbupload = False, and voucherNo_empty    
        return noDataInDF, False, voucherNo_empty   # Data upload failure
        
    except Exception as e:
        print(f'Error: {e}')

        return noDataInDF, False, voucherNo_empty  # Data upload failure
    
#upload_input_values_DB('spsp', '2122', '123456', 'SGCL-PRJ-00235', 'SGCL-PRJ-00285', 'C:\\Users\\Q0037\\Documents\\Malabar\\VoucherVerificationProcess\\InputFolder\\Nehamathew@gmail.com\\input.xlsx','mysql+pymysql://root:Password.123@127.0.0.1:3306/malabarprocessdb','nibiyamonas')


def read_data_from_database(engine_str):

    """
    This is used to fetch data from database having status columns New
    then all sorted data is returned based on the legal entity.

    """
    try:

        #engine = create_engine('mysql+pymysql://root:Password.123@127.0.0.1:3306/malabarprocessdb')
        engine = create_engine(configVariables.dbConnectionString)

        query = "SELECT * FROM project_petty_table WHERE status = 'New'"

        # Execute the query and fetch data into a DataFrame
        df = pd.read_sql(query, con=engine)

        df['prefix'] = df['voucher_legalentity'].str.split('_').str[0]
        df = df.sort_values(by='prefix')

        return df.to_dict(orient='records')
    except Exception as e:
        print("An error occurred:", str(e))
        return None
        


#--------------------------------------------------------------------for looping-------------------------------------------------------------------

def update_status_to_DB(status, comment, primary_key, pymysql_conn_str):
    """
    Update the status and comment columns in the database for a given primary key

    Parameters:
    primary_key (str): The primary key value voucher_legalentity.
    status (str): The new status value.
    comment (str): The new comment value.
    """

    try:
        conn = pymysql.connect(host='127.0.0.1',user='root',password='root',database='malabarprocessdb',port=3306)
        #conn = pymysql.connect(pymysql_conn_str)
            #'mysql+pymysql://root:Password.123@127.0.0.1:3306/malabarprocessdb')
        

        cursor = conn.cursor()

        sql = f"UPDATE project_petty_table SET status = %s, comments = %s WHERE voucher_legalentity = %s"
        cursor.execute(sql, (status, comment, primary_key))

        conn.commit()

        cursor.close()
        conn.close()

        print("Update successful.")
        update_status = True
    except Exception as e:
        update_status = str(e)
        print("An error occurred:", str(e))
        # exceptionMsg = 

    return  update_status

#update_status_to_DB('Valid', 'successfull row data','BGJP_159.0')