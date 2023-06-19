

columns_list = ['status', 'comment']
worksheet, column_dict = create_columns_if_not_exists('C:/Users/Q0037/Documents/Malabar/VoucherVerificationProcess/InputFolder/Nehamathew@gmail.com/input.xlsx', columns_list)




def update_extracted_value(filename, criteriacolumnheader, criteriacellvalue, manufacturerNameHeader, manufacturervalue, MRPheader, MRPvalue, Offerpriceheader, offerpricevalue, quantityheader, quantityvalue):
    """
    Updates the excel file with the given values.

    Returns:
        True if the value is updated.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    # ws, statuscolumn = create_column_if_not_exists(ws, statuscolumnheader)
    ws, statuscolumn = create_columns_if_not_exists(ws, [manufacturerNameHeader, MRPheader, Offerpriceheader, quantityheader])
    for col in ws.iter_cols():
        if criteriacolumnheader == col[0].value:
            for row in ws.iter_rows():
                if criteriacellvalue == str(row[col[0].column-1].value):
                    # row[statuscolumn].value = statusvalue
                    row[statuscolumn[manufacturerNameHeader]].value = manufacturervalue
                    row[statuscolumn[MRPheader]].value = MRPvalue
                    row[statuscolumn[Offerpriceheader]].value = offerpricevalue
                    row[statuscolumn[quantityheader]].value = quantityvalue
                    wb.save(filename)
                    return True
